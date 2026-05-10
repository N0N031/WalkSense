from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from docx import Document

# ============================================================================
# ÉTAPE 1 : EXTRAIRE LE DOCX
# ============================================================================

def extract_docx_sections(docx_path):
    """Extrait les sections principales du DOCX V7.3"""
    doc = Document(docx_path)
    sections = {}
    current_section = None
    
    for paragraph in doc.paragraphs:
        text = paragraph.text.strip()
        if text.startswith("PARTIE"):
            current_section = text
            sections[current_section] = []
        elif current_section and text:
            sections[current_section].append(text)
    
    return sections

# ============================================================================
# STYLES GLOBAUX
# ============================================================================

header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
doctrine_fill = PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid")
critical_fill = PatternFill(start_color="FF4757", end_color="FF4757", fill_type="solid")
todo_fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
done_fill = PatternFill(start_color="51CF66", end_color="51CF66", fill_type="solid")
info_fill = PatternFill(start_color="A8E6CF", end_color="A8E6CF", fill_type="solid")

def apply_header(ws, row, columns):
    for col_num, col_title in enumerate(columns, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = col_title
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def set_cell(ws, row, col, value, fill=None, bold=False):
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if fill:
        cell.fill = fill
    if bold:
        cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

# ============================================================================
# CRÉER EXCEL INTÉGRÉ
# ============================================================================

def create_integrated_excel(docx_path, output_path):
    """Crée Excel fusionnant DOCX V7.3 + structure existante"""
    
    wb = Workbook()
    wb.remove(wb.active)
    docx_sections = extract_docx_sections(docx_path)
    
    # ========================================================================
    # SHEET 0 : TABLE OF CONTENTS
    # ========================================================================
    
    ws_toc = wb.create_sheet("00_TABLE_OF_CONTENTS", 0)
    ws_toc.column_dimensions['A'].width = 40
    ws_toc.column_dimensions['B'].width = 15
    ws_toc.column_dimensions['C'].width = 65
    
    title = ws_toc['A1']
    title.value = "GeoSense Ecosystem — Master Reference v2.1 (DOCX V7.3 Integrated)"
    title.font = Font(bold=True, size=14, color="FFFFFF")
    title.fill = critical_fill
    ws_toc.merge_cells('A1:C1')
    
    apply_header(ws_toc, 3, ["SHEET", "TYPE", "DESCRIPTION"])
    
    toc_entries = [
        ("01_DOCTRINE", "Core", "Vision + mission + valeurs (DOCX)"),
        ("02_ARCHITECTURE", "Tech", "Stack + structure modules"),
        ("03_WALKSENSE_MODULES", "Product", "44 fonctionnalités + status"),
        ("04_ROCKSENSE_HARDWARE", "Hardware", "Specs détecteur EM"),
        ("05_ROCKSENSE_FIRMWARE", "Hardware", "C++ firmware specs"),
        ("06_ROCKSENSE_BLE", "Hardware", "BLE protocol UUID + format"),
        ("07_SEASENSE_VISION", "Future", "Sonar marin vision (DOCX)"),
        ("08_MAPSENSE_VISION", "Future", "Cloud + IA vision (DOCX)"),
        ("09_GRID_SYSTEM", "Feature", "Cellules spatiales (DOCX)"),
        ("10_DATABASE_SCHEMA", "Tech", "SQLite tables + colonnes"),
        ("11_API_DESIGN", "Tech", "REST endpoints"),
        ("12_TECH_STACK", "Tech", "Composants + technos"),
        ("13_UX_FLOWS", "UX", "Flows d'utilisation"),
        ("14_UX_SCREENS", "UX", "Écrans mockups"),
        ("15_SECURITY", "Compliance", "Sécurité + RGPD"),
        ("16_BUSINESS_MODEL", "Strategy", "Modèle économique (DOCX)"),
        ("17_ROADMAP_MASTER", "Timeline", "Sprints + milestones"),
        ("18_TECH_DEBT_RISKS", "Risks", "Bugs + tech debt"),
        ("19_KEY_DECISIONS", "Strategy", "Décisions critiques (DOCX)"),
        ("20_RISKS_MITIGATION", "Management", "Risques + mitigations"),
        ("21_MONETIZATION", "Strategy", "Revenue streams + projections"),
        ("22_TEAM_RESOURCES", "People", "Roles + hiring timeline"),
        ("23_ACQUISITION", "Growth", "Cibles + stratégie (DOCX)"),
        ("24_NEXT_STEPS", "Action", "Actions immédiates prioritaires"),
        ("25_GLOSSARY", "Reference", "Glossaire termes clefs"),
        ("26_DOCX_SYNTHESIS", "Reference", "Résumé DOCX V7.3 complet"),
    ]
    
    row = 4
    for sheet_name, type_, desc in toc_entries:
        set_cell(ws_toc, row, 1, sheet_name, bold=True)
        set_cell(ws_toc, row, 2, type_)
        set_cell(ws_toc, row, 3, desc)
        row += 1
    
    # ========================================================================
    # SHEET 1 : DOCTRINE
    # ========================================================================
    
    ws_doctrine = wb.create_sheet("01_DOCTRINE", 1)
    ws_doctrine.column_dimensions['A'].width = 35
    ws_doctrine.column_dimensions['B'].width = 75
    
    title = ws_doctrine['A1']
    title.value = "GeoSense — Doctrine Produit (DOCX V7.3 PARTIE 2)"
    title.font = Font(bold=True, size=13, color="FFFFFF")
    title.fill = doctrine_fill
    ws_doctrine.merge_cells('A1:B1')
    
    apply_header(ws_doctrine, 2, ["ASPECT", "DÉFINITION"])
    
    doctrine_items = [
        ("MISSION", "Transformer la prospection terrain en système d'information maîtrisé, transparent et évolutif"),
        ("VISION 5 ANS", "Infrastructure nationale voir internationale de données géolocalisées, open-source et anonymisée"),
        ("VALEUR CORE #1", "Traçabilité complète (GPS + photo + hash) = preuve d'intention"),
        ("VALEUR CORE #2", "Respect propriétaire (signature + conformité) = légalité"),
        ("VALEUR CORE #3", "Données exploitables = intelligence terrain"),
        ("POSITIONNEMENT", "Pas un détecteur. Pas un GPS. Un SYSTÈME d'organisation prospection."),
        ("VS COMPETITION", "Nokta MI-6 = hardware. Nous = logiciel structuré + données"),
        ("DIFFÉRENCIATION", "Grid System + Group mode + Cloud future + Anonymisation obligatoire"),
        ("DOCTRINE #1", "Utile avant d'être morale"),
        ("DOCTRINE #2", "Adoption avant monétisation"),
        ("DOCTRINE #3", "Habitudes avant réglementation"),
        ("DOCTRINE #4", "Simplicité avant sophistication"),
        ("DOCTRINE #5", "Terrain-first"),
        ("DOCTRINE #6", "Offline-first"),
        ("DOCTRINE #7", "Privacy-first"),
        ("PRINCIPE CLEF", "Le produit visible doit rester simple même si le moteur interne devient complexe"),
    ]
    
    row = 3
    for aspect, definition in doctrine_items:
        set_cell(ws_doctrine, row, 1, aspect, bold=True, fill=doctrine_fill)
        set_cell(ws_doctrine, row, 2, definition)
        row += 1
    
    # ========================================================================
    # SHEET 2 : ARCHITECTURE
    # ========================================================================
    
    ws_arch = wb.create_sheet("02_ARCHITECTURE", 2)
    for col in range(1, 6):
        ws_arch.column_dimensions[get_column_letter(col)].width = 18
    
    apply_header(ws_arch, 1, ["LAYER", "COMPONENT", "TECH", "STATUS", "NOTES"])
    
    arch_items = [
        ("UI", "React Components", "React Native", "✅ V1.0", "Expo SDK 54"),
        ("UI", "Maps", "Google Maps + Mapbox", "✅ V1.0", "Fallback offline"),
        ("UI", "Router", "Expo Router", "✅ V1.0", "Navigation stack"),
        ("State", "State Management", "Zustand + Context", "✅ V1.0", "Minimal overhead"),
        ("Data", "Local Database", "SQLite + Room", "✅ V1.0", "Offline-first"),
        ("Data", "Async Storage", "AsyncStorage", "✅ V1.0", "Small keys"),
        ("Services", "GPS", "Expo Location", "✅ V1.0", "1-2 Hz tracking"),
        ("Services", "Camera", "Expo ImagePicker", "✅ V1.0", "Photo capture"),
        ("Services", "BLE", "React Native BLE Plx", "❌ V2+", "RockSense future"),
        ("Services", "Crypto", "Expo Crypto", "✅ V1.0", "SHA-256 signing"),
        ("Export", "PDF", "PDFKit", "✅ V1.0", "Client-side"),
        ("Export", "JSON", "JSON.stringify", "✅ V1.0", "No server"),
        ("Cloud", "Backend", "Node.js + Express", "❌ W44+", "Future TypeScript"),
        ("Cloud", "Database", "PostgreSQL", "❌ W44+", "EU hosted"),
        ("Cloud", "Auth", "JWT + bcrypt", "❌ W44+", "Stateless"),
    ]
    
    row = 2
    for layer, component, tech, status, notes in arch_items:
        ws_arch.cell(row=row, column=1).value = layer
        ws_arch.cell(row=row, column=2).value = component
        ws_arch.cell(row=row, column=3).value = tech
        ws_arch.cell(row=row, column=4).value = status
        ws_arch.cell(row=row, column=5).value = notes
        row += 1
    
    # ========================================================================
    # SHEET 3 : WALKSENSE MODULES (44 items)
    # ========================================================================
    
    ws_modules = wb.create_sheet("03_WALKSENSE_MODULES", 3)
    for col in range(1, 8):
        ws_modules.column_dimensions[get_column_letter(col)].width = 16
    
    apply_header(ws_modules, 1, ["MODULE", "BRANCHE", "STATUS", "PRIORITY", "DEPS", "OWNER", "ETA"])
    
    modules_data = [
        ("GPS Live", "WalkSense", "✅ V1.0", "CRITICAL", "Expo Location", "Arnaud", "W1"),
        ("Session CRUD", "WalkSense", "✅ V1.0", "CRITICAL", "SessionManager", "Arnaud", "W2"),
        ("Pre-Session Form", "WalkSense", "✅ V1.0", "CRITICAL", "Validation", "Arnaud", "W3"),
        ("Marker Creation", "WalkSense", "✅ V1.0", "HIGH", "GPS + DB", "Arnaud", "W4"),
        ("Photo Capture", "WalkSense", "✅ V1.0", "HIGH", "Camera API", "Arnaud", "W5"),
        ("Photo Scale", "WalkSense", "✅ V1.0", "MEDIUM", "Metadata", "Arnaud", "W6"),
        ("Classification", "WalkSense", "✅ V1.0", "HIGH", "Enum", "Arnaud", "W7"),
        ("Notes System", "WalkSense", "✅ V1.0", "MEDIUM", "DB", "Arnaud", "W8"),
        ("Map Display (Google)", "WalkSense", "✅ V1.0", "CRITICAL", "Maps API", "Arnaud", "W9"),
        ("Map Display (OSM)", "WalkSense", "✅ V1.0", "HIGH", "Mapbox", "Arnaud", "W10"),
        ("Map Display (IGN)", "WalkSense", "✅ V1.0", "HIGH", "IGN WMS", "Arnaud", "W11"),
        ("Coverage Mapping", "WalkSense", "✅ V1.0", "MEDIUM", "Geometry", "Arnaud", "W12"),
        ("Hash SHA-256", "WalkSense", "✅ V1.0", "CRITICAL", "Crypto", "Arnaud", "W13"),
        ("Session Lock", "WalkSense", "✅ V1.0", "CRITICAL", "DB constraint", "Arnaud", "W14"),
        ("Export JSON", "WalkSense", "✅ V1.0", "HIGH", "JSON", "Arnaud", "W15"),
        ("Export GPX", "WalkSense", "✅ V1.0", "HIGH", "GPX format", "Arnaud", "W16"),
        ("Export PDF", "WalkSense", "✅ V1.0", "MEDIUM", "PDFKit", "Arnaud", "W17"),
        ("Session Historique", "WalkSense", "✅ V1.0", "HIGH", "DB query", "Arnaud", "W18"),
        ("Replay Mode", "WalkSense", "✅ V1.0", "MEDIUM", "Animation", "Arnaud", "W19"),
        ("Geolocation", "WalkSense", "✅ V1.0", "MEDIUM", "Reverse geocoding", "Arnaud", "W20"),
        ("Notifications GPS", "WalkSense", "🟡 TODO", "MEDIUM", "BLE + GPS", "Arnaud", "W21"),
        ("Notifications Marker", "WalkSense", "🟡 TODO", "MEDIUM", "Event listener", "Arnaud", "W22"),
        ("Offline Mode", "WalkSense", "✅ V1.0", "CRITICAL", "SQLite sync", "Arnaud", "W23"),
        ("Battery Optimization", "WalkSense", "🟡 TODO", "HIGH", "OS settings", "Arnaud", "W24"),
        ("Confidentiality Privé", "WalkSense", "✅ V1.0", "HIGH", "Access control", "Arnaud", "W25"),
        ("Confidentiality Groupe", "WalkSense", "✅ V1.0", "HIGH", "Sharing logic", "Arnaud", "W26"),
        ("Confidentiality Fantôme", "WalkSense", "🟡 TODO", "MEDIUM", "Anonymization", "Arnaud", "W27"),
        ("User Profile", "WalkSense", "✅ V1.0", "MEDIUM", "User model", "Arnaud", "W28"),
        ("Group Mode", "WalkSense", "🟡 TODO", "HIGH", "Sync", "Arnaud", "W29"),
        ("Group Chat", "WalkSense", "❌ V2+", "LOW", "WebSocket", "Arnaud", "W30"),
        ("Coverage Analysis", "WalkSense", "🟡 TODO", "MEDIUM", "Geometric algo", "Arnaud", "W31"),
        ("Grid System UI", "WalkSense", "🟡 TODO", "HIGH", "Canvas/SVG", "Arnaud", "W32"),
        ("Grid System Logic", "WalkSense", "🟡 TODO", "CRITICAL", "Math triangulation", "Arnaud", "W33"),
        ("Event Classification", "WalkSense", "✅ V1.0", "HIGH", "Enum + DB", "Arnaud", "W34"),
        ("Camera Permission", "WalkSense", "✅ V1.0", "CRITICAL", "iOS + Android", "Arnaud", "W35"),
        ("GPS Permission", "WalkSense", "✅ V1.0", "CRITICAL", "Location", "Arnaud", "W36"),
        ("Storage Permission", "WalkSense", "✅ V1.0", "HIGH", "File system", "Arnaud", "W37"),
        ("BLE Scan (future)", "WalkSense", "❌ V2+", "HIGH", "React Native BLE", "Arnaud", "W38"),
        ("Signature Propriétaire", "WalkSense", "✅ V1.0", "CRITICAL", "Crypto + PDF", "Arnaud", "W39"),
        ("Dark Mode", "WalkSense", "🟡 TODO", "LOW", "Theme system", "Arnaud", "W40"),
        ("Multi-language", "WalkSense", "❌ V2+", "LOW", "i18n", "Arnaud", "W41"),
        ("Analytics", "WalkSense", "❌ V2+", "LOW", "Firebase/Sentry", "Arnaud", "W42"),
        ("Push Notifications", "WalkSense", "❌ V2+", "MEDIUM", "Firebase Cloud Messaging", "Arnaud", "W43"),
        ("Sync Cloud", "WalkSense", "❌ V2+", "CRITICAL", "Backend API", "Arnaud", "W44"),
    ]
    
    row = 2
    for module, branch, status, priority, deps, owner, eta in modules_data:
        ws_modules.cell(row=row, column=1).value = module
        ws_modules.cell(row=row, column=2).value = branch
        ws_modules.cell(row=row, column=3).value = status
        ws_modules.cell(row=row, column=4).value = priority
        ws_modules.cell(row=row, column=5).value = deps
        ws_modules.cell(row=row, column=6).value = owner
        ws_modules.cell(row=row, column=7).value = eta
        
        if "✅" in status:
            ws_modules.cell(row=row, column=3).fill = done_fill
        elif "🟡" in status:
            ws_modules.cell(row=row, column=3).fill = todo_fill
        elif "❌" in status:
            ws_modules.cell(row=row, column=3).fill = critical_fill
        
        row += 1
    
    # ========================================================================
    # SHEET 4 : ROCKSENSE HARDWARE
    # ========================================================================
    
    ws_rock_hw = wb.create_sheet("04_ROCKSENSE_HARDWARE", 4)
    for col in range(1, 7):
        ws_rock_hw.column_dimensions[get_column_letter(col)].width = 18
    
    title = ws_rock_hw['A1']
    title.value = "RockSense — Hardware BOM (DOCX PARTIE 16)"
    title.font = Font(bold=True, size=12, color="FFFFFF")
    title.fill = critical_fill
    ws_rock_hw.merge_cells('A1:F1')
    
    apply_header(ws_rock_hw, 2, ["COMPONENT", "VERSION", "STATUS", "SPECS", "COST", "ETA"])
    
    rock_hw = [
        ("Boîtier Main", "V16", "❌ V2+", "IP67 renforcé, 400g", "~80€", "W16"),
        ("Capteur ERT", "V16-4ch", "❌ V2+", "4 canaux 0-99, 3-20kHz", "~150€", "W16"),
        ("Batterie Li-ion", "5000mAh", "❌ V2+", "8h autonomie USB-C", "~25€", "W16"),
        ("Écran OLED", "2.4 pouces", "❌ V2+", "240x320 couleur", "~15€", "W16"),
        ("GPS Module", "u-blox NEO-M8N", "❌ V2+", "Précision 2.5m, 10 Hz", "~30€", "W16"),
        ("BLE Module", "nRF52832", "❌ V2+", "BLE 5.0, 100m", "~10€", "W16"),
        ("Microcontrôleur", "ESP32-S3", "❌ V2+", "Dual core, WiFi/BLE", "~5€", "W16"),
        ("Pinpointer", "V16-pin", "❌ V2+", "Détection 30cm, vibration", "~80€", "W17"),
        ("Boîtier Transport", "Custom case", "❌ V2+", "Mousse + compartiments", "~30€", "W17"),
        ("Manuel Utilisateur", "FR/EN", "❌ V2+", "20 pages + QR codes", "~2€", "W18"),
        ("Certification CE", "Compliance", "❌ V2+", "EMC + RoHS", "External lab", "W20"),
    ]
    
    row = 3
    for component, version, status, specs, cost, eta in rock_hw:
        ws_rock_hw.cell(row=row, column=1).value = component
        ws_rock_hw.cell(row=row, column=2).value = version
        ws_rock_hw.cell(row=row, column=3).value = status
        ws_rock_hw.cell(row=row, column=4).value = specs
        ws_rock_hw.cell(row=row, column=5).value = cost
        ws_rock_hw.cell(row=row, column=6).value = eta
        ws_rock_hw.cell(row=row, column=3).fill = critical_fill
        row += 1
    
    # ========================================================================
    # SHEET 5 : ROCKSENSE FIRMWARE
    # ========================================================================
    
    ws_rock_fw = wb.create_sheet("05_ROCKSENSE_FIRMWARE", 5)
    for col in range(1, 8):
        ws_rock_fw.column_dimensions[get_column_letter(col)].width = 16
    
    apply_header(ws_rock_fw, 1, ["FONCTION", "LANGAGE", "STATUS", "PRIORITY", "DESCRIPTION", "DEPS", "ETA"])
    
    rock_fw = [
        ("Boot Sequence", "C++", "❌ V2+", "CRITICAL", "Init capteurs + BLE", "ESP32 SDK", "W16"),
        ("ERT Reading", "C++", "❌ V2+", "CRITICAL", "Échantillonnage 4 canaux", "Driver Nokta", "W16"),
        ("GPS Position", "C++", "❌ V2+", "CRITICAL", "Parse NMEA, buffer circulaire", "u-blox lib", "W16"),
        ("BLE Broadcasting", "C++", "❌ V2+", "HIGH", "Beacon format custom JSON", "nRF SDK", "W17"),
        ("Data Buffering", "C++", "❌ V2+", "HIGH", "FIFO 1000 samples, SD fallback", "FAT FS", "W17"),
        ("Battery Monitor", "C++", "❌ V2+", "MEDIUM", "ADC + voltage cutoff", "ESP32 ADC", "W17"),
        ("OLED Display", "C++", "❌ V2+", "MEDIUM", "Real-time graph + menu", "Display lib", "W18"),
        ("OTA Update", "C++", "❌ V2+", "HIGH", "Firmware update OTA via BLE", "Arduino OTA", "W18"),
    ]
    
    row = 2
    for fonction, langage, status, priority, description, deps, eta in rock_fw:
        ws_rock_fw.cell(row=row, column=1).value = fonction
        ws_rock_fw.cell(row=row, column=2).value = langage
        ws_rock_fw.cell(row=row, column=3).value = status
        ws_rock_fw.cell(row=row, column=4).value = priority
        ws_rock_fw.cell(row=row, column=5).value = description
        ws_rock_fw.cell(row=row, column=6).value = deps
        ws_rock_fw.cell(row=row, column=7).value = eta
        row += 1
    
    # ========================================================================
    # SHEET 6 : ROCKSENSE BLE PROTOCOL
    # ========================================================================
    
    ws_ble = wb.create_sheet("06_ROCKSENSE_BLE", 6)
    ws_ble.column_dimensions['A'].width = 35
    ws_ble.column_dimensions['B'].width = 60
    
    apply_header(ws_ble, 1, ["PARAMETER", "VALUE"])
    
    ble_data = [
        ("Service UUID", "550e8400-e29b-41d4-a716-446655440000"),
        ("Characteristic Data", "550e8400-e29b-41d4-a716-446655440001 (notify)"),
        ("Characteristic Status", "550e8400-e29b-41d4-a716-446655440002 (read)"),
        ("Data Format", "JSON: {ts, lat, lon, ch1, ch2, ch3, ch4, rssi}"),
        ("Broadcast Interval", "1 second"),
        ("MTU", "512 bytes (fragmentation si > 512)"),
        ("Range", "100m line-of-sight (nRF52832)"),
        ("Encryption", "BLE Level 2 (pairing + bonding)"),
        ("Reconnect Logic", "Exponential backoff max 30s"),
        ("Status Fields", "{battery %, firmware version, samples buffered}"),
    ]
    
    row = 2
    for param, value in ble_data:
        ws_ble.cell(row=row, column=1).value = param
        ws_ble.cell(row=row, column=2).value = value
        row += 1
    
    # ========================================================================
    # SHEET 7 : SEASENSE VISION
    # ========================================================================
    
    ws_sea = wb.create_sheet("07_SEASENSE_VISION", 7)
    ws_sea.column_dimensions['A'].width = 40
    ws_sea.column_dimensions['B'].width = 60
    
    title = ws_sea['A1']
    title.value = "SeaSense — Sonar marin (DOCX PARTIE 7)"
    title.font = Font(bold=True, size=12, color="FFFFFF")
    title.fill = critical_fill
    ws_sea.merge_cells('A1:B1')
    
    apply_header(ws_sea, 2, ["ASPECT", "DÉFINITION"])
    
    sea_data = [
        ("VISION", "Sonar + ROV → cartographie fonds marins + gisements"),
        ("OBJECTIF", "Compléter RockSense avec données marines"),
        ("DÉCLENCHEUR", "RockSense stable (W32) + funding"),
        ("TIMING", "Week 32-48 R&D → Week 50+ alpha"),
        ("HARDWARE CONCEPT", "À DÉFINIR : Sonar portable ? Cartographie cours d'eau ? Scan littoral ?"),
        ("HARDWARE SPECS", "À DÉFINIR : Sonar écholoticateur ? Drone aquatique ? Capteur submersible ?"),
        ("SOFTWARE", "À DÉFINIR : Integration GPS / BLE / Data"),
        ("MARCHÉ", "À DÉFINIR : Archéologie aquatique ? Exploration littoral ? Pêche ? Sauvetage ?"),
        ("BUDGET ESTIMÉ", "À DÉFINIR : 5k-20k€ selon complexité"),
        ("PRIORITY", "🔵 FUTURE : Dépend succès WalkSense"),
    ]
    
    row = 3
    for aspect, definition in sea_data:
        set_cell(ws_sea, row, 1, aspect, bold=True, fill=info_fill)
        set_cell(ws_sea, row, 2, definition)
        row += 1
    
    # ========================================================================
    # SHEET 8 : MAPSENSE VISION
    # ========================================================================
    
    ws_map = wb.create_sheet("08_MAPSENSE_VISION", 8)
    ws_map.column_dimensions['A'].width = 40
    ws_map.column_dimensions['B'].width = 60
    
    title = ws_map['A1']
    title.value = "MapSense — Cloud + GeoFusion + IA (DOCX PARTIE 8)"
    title.font = Font(bold=True, size=12, color="FFFFFF")
    title.fill = critical_fill
    ws_map.merge_cells('A1:B1')
    
    apply_header(ws_map, 2, ["ASPECT", "DÉFINITION"])
    
    map_data = [
        ("VISION", "Plateforme cloud : fusion WalkSense+RockSense+SeaSense+IA"),
        ("OBJECTIF", "Base collective anonyme, patterns spatiaux, HiddenMap IA"),
        ("DÉCLENCHEUR", "Mode groupe stable (W36) + financement"),
        ("TIMING", "Week 37-44 MVP → Week 45-60+ IA"),
        ("DATA COLLECTION", "Sessions WalkSense → optionnel upload cloud (user consent)"),
        ("ANONYMISATION", "Détail : suppression identité utilisateur, conservation GPS/type/classe"),
        ("STORAGE", "PostgreSQL on Hetzner (EU) ou AWS (RGPD compliant)"),
        ("API REST", "GET /cells/{id}, POST /session, GET /heatmap/{region}"),
        ("VISUALIZATION", "Web dashboard : heatmap densité, timeline, filtres géo/type"),
        ("ML FUTURE", "IA classification image, prédiction zones riches, clustering spatial"),
        ("REVENUE", "B2B : data access pour archéologues/institutions (€1000/an)"),
        ("COMPLIANCE", "RGPD : opt-in, data anonymized, user can request deletion"),
    ]
    
    row = 3
    for aspect, definition in map_data:
        set_cell(ws_map, row, 1, aspect, bold=True, fill=info_fill)
        set_cell(ws_map, row, 2, definition)
        row += 1
    
    # ========================================================================
    # SHEET 9 : GRID SYSTEM
    # ========================================================================
    
    ws_grid = wb.create_sheet("09_GRID_SYSTEM", 9)
    ws_grid.column_dimensions['A'].width = 30
    ws_grid.column_dimensions['B'].width = 65
    
    title = ws_grid['A1']
    title.value = "Grid System — Système cellules (DOCX PARTIE 11)"
    title.font = Font(bold=True, size=12, color="FFFFFF")
    title.fill = critical_fill
    ws_grid.merge_cells('A1:B1')
    
    apply_header(ws_grid, 2, ["ÉLÉMENT", "DESCRIPTION"])
    
    grid_data = [
        ("CONCEPT", "Core 4 piquets = 4 points GPS connus → création grille virtuelle"),
        ("DÉFINITION", "Rectangle 4x4 ou 8x8 ou 16x16 cellules (configurable)"),
        ("NOMMAGE", "Cellule A1, A2, ..., P16 = adressable par prospecteur"),
        ("POSITION", "GPS brut → projection sur grille → cellule estimée"),
        ("PRÉCISION", "Dépend : taille grille + précision GPS (±5m standard)"),
        ("MATH", "Transformation affine : (lat,lon) → (grid_x, grid_y) → cellule"),
        ("IMPLÉMENTATION", "Fonction simple trigonométrie en TypeScript"),
        ("USE CASE", "Prospecteur marque zone riche → peut dire 'C4, C5, D5' vs GPS exact"),
        ("AVANTAGE", "Communication simple, pas de doublons, partage groupe facile"),
        ("UI DISPLAY", "Overlay grille sur carte + affichage cellule courante"),
        ("DATABASE", "Table cells : id, grid_id, name, markers_count, coverage %"),
        ("FUTURE", "Triangulation multi-piquet pour améliorer précision"),
    ]
    
    row = 3
    for element, description in grid_data:
        set_cell(ws_grid, row, 1, element, bold=True, fill=info_fill)
        set_cell(ws_grid, row, 2, description)
        row += 1
    
    # ========================================================================
    # SHEET 10 : DATABASE SCHEMA
    # ========================================================================
    
    ws_db = wb.create_sheet("10_DATABASE_SCHEMA", 10)
    for col in range(1, 8):
        ws_db.column_dimensions[get_column_letter(col)].width = 15
    
    apply_header(ws_db, 1, ["TABLE", "COLONNE", "TYPE", "PRIMAIRE", "FK", "INDEX", "DESCRIPTION"])
    
    db_schema = [
        ("users", "id", "String", "PK", "—", "yes", "UUID"),
        ("users", "name", "String", "—", "—", "—", "Nom"),
        ("users", "email", "String", "UNIQUE", "—", "yes", "Email login"),
        ("users", "created_at", "Long", "—", "—", "—", "Timestamp"),
        ("sessions", "id", "String", "PK", "—", "yes", "UUID"),
        ("sessions", "user_id", "String", "—", "users.id", "yes", "Owner"),
        ("sessions", "title", "String", "—", "—", "—", "Nom session"),
        ("sessions", "start_time", "Long", "—", "—", "—", "Début"),
        ("sessions", "end_time", "Long?", "—", "—", "—", "Fin (null=ongoing)"),
        ("sessions", "status", "Enum", "—", "—", "yes", "ONGOING/PAUSED/CLOSED"),
        ("sessions", "location", "String", "—", "—", "—", "Commune + dept"),
        ("sessions", "grid_id", "String", "—", "grids.id", "—", "FK grille"),
        ("sessions", "hash", "String", "—", "—", "yes", "SHA-256 integrity"),
        ("sessions", "is_locked", "Boolean", "—", "—", "—", "Locked after close"),
        ("gps_points", "id", "String", "PK", "—", "yes", "UUID"),
        ("gps_points", "session_id", "String", "—", "sessions.id", "yes", "FK"),
        ("gps_points", "lat", "Double", "—", "—", "—", "Latitude"),
        ("gps_points", "lon", "Double", "—", "—", "—", "Longitude"),
        ("gps_points", "accuracy", "Float", "—", "—", "—", "Mètres"),
        ("gps_points", "timestamp", "Long", "—", "—", "yes", "Quand collecté"),
        ("gps_points", "speed", "Float?", "—", "—", "—", "Vitesse m/s"),
        ("markers", "id", "String", "PK", "—", "yes", "UUID"),
        ("markers", "session_id", "String", "—", "sessions.id", "yes", "FK"),
        ("markers", "lat", "Double", "—", "—", "yes", "Position"),
        ("markers", "lon", "Double", "—", "—", "yes", "Position"),
        ("markers", "classification", "Enum", "—", "—", "yes", "DECHET/MODERNE/..."),
        ("markers", "notes", "String?", "—", "—", "—", "Notes libres"),
        ("markers", "cell_id", "String", "—", "cells.id", "—", "FK si grid"),
        ("photos", "id", "String", "PK", "—", "yes", "UUID"),
        ("photos", "marker_id", "String", "—", "markers.id", "yes", "FK"),
        ("photos", "file_path", "String", "—", "—", "—", "Local path"),
        ("photos", "scale_reference", "String?", "—", "—", "—", "Échelle"),
        ("photos", "taken_at", "Long", "—", "—", "yes", "Timestamp"),
        ("grids", "id", "String", "PK", "—", "yes", "UUID"),
        ("grids", "session_id", "String", "—", "sessions.id", "yes", "FK"),
        ("grids", "anchor1_lat", "Double", "—", "—", "—", "Piquet 1"),
        ("grids", "anchor1_lon", "Double", "—", "—", "—", "Piquet 1"),
        ("grids", "grid_size", "Int", "—", "—", "—", "4/8/16"),
        ("cells", "id", "String", "PK", "—", "yes", "UUID"),
        ("cells", "grid_id", "String", "—", "grids.id", "yes", "FK"),
        ("cells", "name", "String", "—", "—", "yes", "A1, B2, ..."),
        ("cells", "markers_count", "Int", "—", "—", "—", "Nb marqueurs"),
        ("cells", "coverage_percent", "Float", "—", "—", "—", "% couvert"),
        ("events", "id", "String", "PK", "—", "yes", "UUID"),
        ("events", "session_id", "String", "—", "sessions.id", "yes", "FK"),
        ("events", "type", "Enum", "—", "—", "yes", "MARKER/BLE/..."),
        ("events", "lat", "Double?", "—", "—", "—", "Position"),
        ("events", "lon", "Double?", "—", "—", "—", "Position"),
        ("events", "created_at", "Long", "—", "—", "yes", "Timestamp"),
        ("ble_devices", "id", "String", "PK", "—", "yes", "UUID"),
        ("ble_devices", "session_id", "String", "—", "sessions.id", "yes", "FK"),
        ("ble_devices", "device_name", "String", "—", "—", "—", "RockSense name"),
        ("ble_devices", "mac_address", "String", "UNIQUE", "—", "yes", "BLE address"),
        ("ble_devices", "battery", "Int", "—", "—", "—", "% batterie"),
        ("session_exports", "id", "String", "PK", "—", "yes", "UUID"),
        ("session_exports", "session_id", "String", "—", "sessions.id", "yes", "FK"),
        ("session_exports", "format", "Enum", "—", "—", "—", "JSON/GPX/PDF"),
        ("session_exports", "file_path", "String", "—", "—", "—", "Chemin"),
        ("session_exports", "created_at", "Long", "—", "—", "yes", "Timestamp"),
    ]
    
    row = 2
    for table, colonne, type_, primary, fk, index, description in db_schema:
        ws_db.cell(row=row, column=1).value = table
        ws_db.cell(row=row, column=2).value = colonne
        ws_db.cell(row=row, column=3).value = type_
        ws_db.cell(row=row, column=4).value = primary
        ws_db.cell(row=row, column=5).value = fk
        ws_db.cell(row=row, column=6).value = index
        ws_db.cell(row=row, column=7).value = description
        row += 1
    
    # ========================================================================
    # SHEET 11 : API DESIGN
    # ========================================================================
    
    ws_api = wb.create_sheet("11_API_DESIGN", 11)
    for col in range(1, 6):
        ws_api.column_dimensions[get_column_letter(col)].width = 18
    
    apply_header(ws_api, 1, ["ENDPOINT", "METHOD", "REQUEST", "RESPONSE", "DESCRIPTION"])
    
    api_data = [
        ("/auth/register", "POST", "name, email, password", "token, user_id", "Inscription"),
        ("/auth/login", "POST", "email, password", "token, expires_at", "Connexion"),
        ("/sessions", "POST", "title, location, user_id", "session_id, hash", "Créer session"),
        ("/sessions/{id}", "GET", "—", "session_data + markers", "Récupérer session"),
        ("/sessions/{id}", "PATCH", "status (CLOSED)", "session_data + hash", "Clôturer session"),
        ("/sessions/{id}/markers", "POST", "lat, lon, class, photo", "marker_id", "Ajouter marqueur"),
        ("/sessions/{id}/markers", "GET", "—", "[markers]", "Lister marqueurs"),
        ("/sessions/{id}/export", "POST", "format (json/gpx/pdf)", "file_url", "Exporter session"),
        ("/grids", "POST", "session_id, anchors[4]", "grid_id", "Créer grille"),
        ("/grids/{id}/cells", "GET", "—", "[cells with data]", "Lister cellules"),
        ("/ble/devices", "POST", "session_id, mac, data", "device_id", "Register RockSense"),
        ("/ble/devices/{id}/data", "GET", "—", "[ert_samples + gps]", "Récupérer données"),
        ("/sync", "POST", "session_data (full)", "sync_id, conflicts", "Sync bidirectionnelle"),
        ("/maps/heatmap", "GET", "region, type", "heatmap_geojson", "Carte agrégée anonyme"),
        ("DELETE /sessions/{id}", "DELETE", "—", "success", "Supprimer session (RGPD)"),
    ]
    
    row = 2
    for endpoint, method, request, response, description in api_data:
        ws_api.cell(row=row, column=1).value = endpoint
        ws_api.cell(row=row, column=2).value = method
        ws_api.cell(row=row, column=3).value = request
        ws_api.cell(row=row, column=4).value = response
        ws_api.cell(row=row, column=5).value = description
        row += 1
    
    # ========================================================================
    # SHEET 12 : TECH STACK
    # ========================================================================
    
    ws_stack = wb.create_sheet("12_TECH_STACK", 12)
    for col in range(1, 4):
        ws_stack.column_dimensions[get_column_letter(col)].width = 25
    
    apply_header(ws_stack, 1, ["COMPOSANT", "TECHNOLOGIE", "RAISON / DÉTAILS"])
    
    stack_data = [
        ("Mobile Framework", "React Native (Expo)", "Cross-platform, rapid dev, hot reload"),
        ("Language", "TypeScript", "Type safety, IDE support, error prevention"),
        ("Database Local", "SQLite + Room", "Offline-first, performant, standard industrie"),
        ("State Management", "Zustand / Context API", "Minimal overhead, good for offline"),
        ("Maps", "Google Maps + Mapbox + IGN", "Multiple fallbacks, offline capable"),
        ("Camera", "Expo ImagePicker", "Simple, permissions handling OK"),
        ("GPS", "Expo Location", "Good accuracy, battery optimization possible"),
        ("BLE", "React Native BLE Plx", "Cross-platform, real-time, stable"),
        ("Crypto", "TweetNaCl.js (crypto_sign)", "SHA-256, signing, lightweight"),
        ("Export", "PDFKit (JS) + json-stringify", "Client-side, no backend needed"),
        ("Testing", "Jest + React Native Testing Lib", "Unit + integration tests"),
        ("CI/CD", "GitHub Actions", "Automated build, EAS Deploy"),
        ("Backend (Future)", "Node.js + Express", "TypeScript, same lang as mobile"),
        ("Backend DB", "PostgreSQL on Hetzner", "ACID compliance, RGPD EU data center"),
        ("Backend Auth", "JWT + bcrypt", "Stateless, secure password hashing"),
        ("API", "REST + JSON", "Simple, well-understood, easy versioning"),
        ("Real-time (Future)", "WebSocket via Socket.io", "Group sync, live updates"),
        ("Cloud Storage", "S3 compatible (Hetzner)", "Cost-effective, EU-hosted"),
        ("Monitoring", "Sentry + LogRocket", "Error tracking + session replay"),
        ("Analytics", "PostHog", "Privacy-first, self-hosted option"),
    ]
    
    row = 2
    for composant, technologie, raison in stack_data:
        ws_stack.cell(row=row, column=1).value = composant
        ws_stack.cell(row=row, column=2).value = technologie
        ws_stack.cell(row=row, column=3).value = raison
        row += 1
    
    # ========================================================================
    # SHEET 13 : UX FLOWS
    # ========================================================================
    
    ws_ux_flows = wb.create_sheet("13_UX_FLOWS", 13)
    ws_ux_flows.column_dimensions['A'].width = 30
    ws_ux_flows.column_dimensions['B'].width = 65
    
    apply_header(ws_ux_flows, 1, ["FLOW NAME", "DESCRIPTION"])
    
    ux_flows = [
        ("ONBOARDING", "Screen 1: Welcome → Screen 2: Passcode → Screen 3: Permissions → Screen 4: Tutorial"),
        ("NEW SESSION", "Tap 'Nouvelle session' → Form → Signature propriétaire → Start → GPS live"),
        ("TRACKING", "Carte 8 couches + position live + compteur distance + trace temps réel"),
        ("MARKER CREATION", "Long press OR tap bouton → Photo → Classification → Notes → Save"),
        ("SESSION PAUSE", "Tap pause → compteur visible → Tap resume → continue"),
        ("SESSION CLOSE", "Tap 'Terminer' → confirmation → hash calculé → locked → résumé"),
        ("REVIEW SESSION", "Historique → Tap session → Carte replay + timeline → Tap marqueur"),
        ("EXPORT", "Session détail → Choix format (JSON/GPX/PDF) → Share ou Save"),
        ("GROUP MODE", "Créer groupe → Invite via code → Sessions visibles → Positions live"),
        ("CONFIDENTIALITY", "Settings → Tap mode (Privé/Flouté/Groupe/Fantôme) → Save"),
        ("NOTIFICATIONS", "GPS warning (orange) → Marker confirm (haptic) → Battery low (rouge)"),
        ("REPLAY", "Session → Tap 'Rejouer' → Play → Vitesse 1-10x → Pause/step"),
    ]
    
    row = 2
    for flow_name, description in ux_flows:
        ws_ux_flows.cell(row=row, column=1).value = flow_name
        ws_ux_flows.cell(row=row, column=2).value = description
        row += 1
    
    # ========================================================================
    # SHEET 14 : UX SCREENS
    # ========================================================================
    
    ws_ux_screens = wb.create_sheet("14_UX_SCREENS", 14)
    ws_ux_screens.column_dimensions['A'].width = 30
    ws_ux_screens.column_dimensions['B'].width = 50
    
    apply_header(ws_ux_screens, 1, ["ÉCRAN", "DESCRIPTION"])
    
    ux_screens = [
        ("Splash Screen", "GeoSense logo + fade-in (2s)"),
        ("Login Screen", "Email input → Password input → Login button → Sign up link"),
        ("Sign Up Screen", "Name input → Email → Password (strength meter) → Terms checkbox"),
        ("Onboarding #1", "Welcome message + logo → 'Démarrez votre 1e session'"),
        ("Onboarding #2", "Passcode setup (4-6 digits) → Repeat → Save"),
        ("Onboarding #3", "Permissions request : GPS (req) → Camera (req) → Storage (req)"),
        ("Onboarding #4", "Tutorial : 'Mode privé = invisible' → Done button"),
        ("Dashboard", "3 sessions récentes → Button 'Nouvelle' → Stats (distance, marqueurs)"),
        ("New Session Form", "TextInput Titre → Motif → Location dropdown → Checkbox → Start"),
        ("Main Tracking", "Header (title + elapsed) → Map → Bottom buttons (Ajouter/Pause/...)"),
        ("Map Layer Selector", "Overlay buttons : Google / Satellite / OSM / IGN / Cassini"),
        ("Add Marker Modal", "Camera → Photo preview + scale → Classification dropdown → Notes"),
        ("Session Pause Screen", "Message 'Pausée' + Timer + Resume button"),
        ("Session End Screen", "Summary : distance / duration / markers / photos → Share/Export"),
        ("Session History", "List sessions + pull-to-refresh → Tap to detail"),
        ("Session Detail", "Header + trace map + events timeline + markers table + export"),
        ("Replay Screen", "Fullscreen map → Play/Pause → Speed slider (1-10x) → Timeline scrubber"),
        ("Export Modal", "Format selection (JSON/GPX/PDF) → File size → Share/Save"),
        ("Settings Screen", "Confidentiality mode (radio) → Language → Dark mode → About → Logout"),
        ("Profile Screen", "Name + avatar + email + stats (distance, markers, sessions) → Delete account"),
        ("Group Management", "Create button → List groups → Members → Leave → Invite code"),
    ]
    
    row = 2
    for ecran, description in ux_screens:
        ws_ux_screens.cell(row=row, column=1).value = ecran
        ws_ux_screens.cell(row=row, column=2).value = description
        row += 1
    
    # ========================================================================
    # SHEET 15 : SECURITY & COMPLIANCE
    # ========================================================================
    
    ws_security = wb.create_sheet("15_SECURITY", 15)
    ws_security.column_dimensions['A'].width = 35
    ws_security.column_dimensions['B'].width = 65
    
    apply_header(ws_security, 1, ["ASPECT", "DÉTAIL"])
    
    security_data = [
        ("RGPD Compliant", "✅ Data stored locally. Opt-in upload only. User can request deletion."),
        ("Data Encryption", "✅ Local: SQLite encryption ready (SQLCipher). Transport: HTTPS only."),
        ("User Auth", "✅ Password hashing (bcrypt). Passcode app-level. Token-based API (future)."),
        ("Session Integrity", "✅ Hash SHA-256 on close. Immutable after lock. Timestamp proof."),
        ("Permission System", "✅ Camera/GPS/Storage explicit consent + OS-level permissions."),
        ("Digital Signature", "✅ PDF signature field for propriétaire authorization."),
        ("Data Minimization", "✅ Only collect : position, photo, classification, timestamp."),
        ("Propriétaire Consent", "✅ Pre-session form: checkbox 'J'autorise'. Signature after."),
        ("No Guarantees", "⚠️ App provides zero legal guarantee. User responsible for local laws."),
        ("Future Audits", "🔵 Plan: 3rd party security audit (W40+), GDPR certification (W50+)."),
        ("Logging", "✅ Error logs stored locally. No personal data. Sentry (future, anonymized)."),
        ("Third-party Risk", "⚠️ Google Maps, BLE. Minimize data sent. Use alternatives when possible."),
    ]
    
    row = 2
    for aspect, detail in security_data:
        ws_security.cell(row=row, column=1).value = aspect
        ws_security.cell(row=row, column=2).value = detail
        row += 1
    
    # ========================================================================
    # SHEET 16 : BUSINESS MODEL
    # ========================================================================
    
    ws_biz = wb.create_sheet("16_BUSINESS_MODEL", 16)
    ws_biz.column_dimensions['A'].width = 40
    ws_biz.column_dimensions['B'].width = 65
    
    apply_header(ws_biz, 1, ["KEY", "VALUE"])
    
    biz_data = [
        ("Target Market", "Prospecteurs loisir (France/Europe) + archéologues + explorateurs"),
        ("TAM (Total)", "~50k prospecteurs actifs en Europe"),
        ("SAM (Serviceable)", "~10k prospecteurs en France → 500 users possible"),
        ("SOM (Obtainable) Y1", "50-100 users → Y2 200-300 → Y3 500-1000"),
        ("Revenue Stream #1", "WalkSense Freemium : 0€ (app gratuit, aucun paywall)"),
        ("Revenue Stream #2", "GeoSense+ Premium : €5/mois (cloud sync, backup, advanced export)"),
        ("Revenue Stream #3", "RockSense Hardware : €450-550/unit (hardware kit détecteur+pinpointer)"),
        ("Revenue Stream #4", "MapSense Data API : €1000/an (institutions, research, archaeology)"),
        ("Revenue Stream #5", "Sponsorship : Constructeurs détecteurs (placement, affiliate)"),
        ("CAC", "Google Ads: €5-10 per user | Organic (YouTube): €0"),
        ("LTV Premium", "€60/user (1 year average)"),
        ("LTV Hardware", "€500 (one-time, 10% margin = €50)"),
        ("Breakeven", "Y2-Y3 depending adoption rate"),
        ("Y1 Revenue Projection", "~€14.3k (600 users mixed : free + premium + 10 hardware)"),
        ("Series A Timing", "W44+ if traction"),
        ("Series A Amount", "€200k"),
        ("Burn Rate (Bootstrap)", "Personal salary only (~€2k/month)"),
        ("Runway", "~12 months (personal savings)"),
        ("Exit Strategy", "Acquire by mapping company, archaeology firm, or IPO (unlikely)"),
        ("Funding Strategy", "Bootstrap Y1 → Pre-seed Y2 (€50k) → Series A Y3 (€200k)"),
    ]
    
    row = 2
    for key, value in biz_data:
        ws_biz.cell(row=row, column=1).value = key
        ws_biz.cell(row=row, column=2).value = value
        row += 1
    
    # ========================================================================
    # SHEET 17 : ROADMAP MASTER
    # ========================================================================
    
    ws_roadmap = wb.create_sheet("17_ROADMAP_MASTER", 17)
    for col in range(1, 5):
        ws_roadmap.column_dimensions[get_column_letter(col)].width = 18
    
    apply_header(ws_roadmap, 1, ["WEEKS", "MILESTONE", "DELIVERABLES", "STATUS"])
    
    roadmap_data = [
        ("W1-W8", "Sprint A.1", "Bug fixes, TypeScript, tests", "🔴 IN PROGRESS"),
        ("W9-W16", "Sprint B.1", "GPS live, markers, photos, export", "🟡 TODO"),
        ("W17-W24", "Sprint B.2", "UI/UX, dark mode, notifications", "🟡 TODO"),
        ("W25-W32", "Sprint B.3", "Group mode, sharing, permissions", "🟡 TODO"),
        ("W33-W40", "Sprint C.1", "RockSense proto OR MapSense MVP", "🟡 TODO"),
        ("W41-W48", "Sprint C.2", "RockSense beta / Backend start", "🔵 FUTURE"),
        ("W49-W56", "Sprint D.1", "Cloud API, PostgreSQL, anonymization", "🔵 FUTURE"),
        ("W57-W64", "Sprint D.2", "Web dashboard, heatmap, data explorer", "🔵 FUTURE"),
        ("W65+", "Sprint E.1", "AI classification, scaling, monitor", "🔵 FUTURE"),
    ]
    
    row = 2
    for weeks, milestone, deliverables, status in roadmap_data:
        ws_roadmap.cell(row=row, column=1).value = weeks
        ws_roadmap.cell(row=row, column=2).value = milestone
        ws_roadmap.cell(row=row, column=3).value = deliverables
        ws_roadmap.cell(row=row, column=4).value = status
        row += 1
    
    # ========================================================================
    # SHEET 18 : TECH DEBT & RISKS
    # ========================================================================
    
    ws_debt = wb.create_sheet("18_TECH_DEBT_RISKS", 18)
    for col in range(1, 7):
        ws_debt.column_dimensions[get_column_letter(col)].width = 16
    
    apply_header(ws_debt, 1, ["BUG/DEBT", "SEVERITY", "AFFECTS", "CAUSE", "FIX ETA", "NOTES"])
    
    debt_data = [
        ("GPS drift long sessions", "🟡 MEDIUM", "Tracking", "Accumulation error", "W12", "Kalman filter"),
        ("Photo orientation iOS", "🔴 HIGH", "Photos", "EXIF handling", "W8", "Native picker"),
        ("BLE reconnect delay", "🟡 MEDIUM", "RockSense", "Backoff timing slow", "W35", "Adjust timing"),
        ("Map offline cache", "🟡 MEDIUM", "Map display", "No offline tile cache", "W20", "Mapbox offline"),
        ("Database indexes", "🟡 MEDIUM", "Performance", "Slow queries big sessions", "W25", "Add indexes"),
        ("Memory leak replay", "🔴 HIGH", "Replay mode", "Listener not cleaned", "W19", "Fix useEffect"),
        ("Export PDF generation", "🟡 MEDIUM", "Export", "Slow 1000+ markers", "W22", "Batch processing"),
        ("UI lag grid 16x16", "🟡 MEDIUM", "Grid System", "256 cells render heavy", "W32", "Virtualization+canvas"),
    ]
    
    row = 2
    for bug, sev, affects, cause, eta, notes in debt_data:
        ws_debt.cell(row=row, column=1).value = bug
        ws_debt.cell(row=row, column=2).value = sev
        ws_debt.cell(row=row, column=3).value = affects
        ws_debt.cell(row=row, column=4).value = cause
        ws_debt.cell(row=row, column=5).value = eta
        ws_debt.cell(row=row, column=6).value = notes
        row += 1
    
       # ========================================================================
    # SHEET 19 : KEY DECISIONS
    # ========================================================================
    
    ws_decisions = wb.create_sheet("19_KEY_DECISIONS", 19)
    ws_decisions.column_dimensions['A'].width = 40
    ws_decisions.column_dimensions['B'].width = 65
    
    apply_header(ws_decisions, 1, ["DECISION", "RATIONALE"])
    
    decisions_data = [
        ("React Native NOT Native", "Cross-platform faster, Expo hot reload, less code duplication"),
        ("SQLite NOT Firebase", "Offline-first, RGPD control, no vendor lock-in, offline-native"),
        ("Google Maps default + Mapbox fallback NOT single map", "Multiple fallbacks ensure offline + reliability"),
        ("1-2 Hz GPS tracking NOT 0.5 Hz", "0.5 Hz = too much drift, 5 Hz = battery drain. 1-2 Hz = sweet spot."),
        ("Custom BLE JSON NOT binary", "JSON = human-readable, debuggable, simple to iterate"),
        ("Client-side export (JSON/GPX/PDF) NOT server", "Works offline, no backend dependency, instant"),
        ("Grid System server-side future NOT hardware", "Test software ecosystem first (W32 decision point)."),
        ("Freemium NOT paid-only", "Adoption > revenue early. Premium features later."),
        ("No legal guarantees", "User full responsibility. Explicit RGPD consent. Get lawyer review (W30)."),
        ("Open-source eventually NOT closed", "Build community trust, attract contributors, future exit value"),
        ("Bootstrap solo now, hire later", "Prove concept first (W24), then hire UI/UX (W24), Dev (W32), DevOps (W44)"),
        ("Offline-first architecture", "Terrain means poor connectivity. Sync asynchronous when online."),
        ("SHA-256 hash for integrity", "Proof of data immutability, session locked after close"),
        ("Propriétaire signature in app", "Legal proof of consent, exported with session data"),
        ("No profiling / No ads", "User privacy core value, sustainable via premium + data API"),
    ]
    
    row = 2
    for decision, rationale in decisions_data:
        ws_decisions.cell(row=row, column=1).value = decision
        ws_decisions.cell(row=row, column=2).value = rationale
        row += 1
    
    # ========================================================================
    # SHEET 20 : RISKS & MITIGATION
    # ========================================================================
    
    ws_risks = wb.create_sheet("20_RISKS_MITIGATION", 20)
    for col in range(1, 5):
        ws_risks.column_dimensions[get_column_letter(col)].width = 22
    
    apply_header(ws_risks, 1, ["RISK", "PROBABILITY", "IMPACT", "MITIGATION"])
    
    risks_data = [
        ("GPS precision inadequate", "MEDIUM", "HIGH", "Field test extensively, add Kalman filter, use multiple sources"),
        ("RockSense hardware delayed", "HIGH", "MEDIUM", "Pivot to software-only longer, pre-orders for funding"),
        ("Market saturation by Nokta", "MEDIUM", "MEDIUM", "Focus on software + data value-add, community lock-in"),
        ("RGPD enforcement backlash", "LOW", "HIGH", "Get legal review (W30), explicit consent always, audit"),
        ("Battery drain daily use", "MEDIUM", "MEDIUM", "Optimize GPS freq, add power mode, background task"),
        ("User adoption slow", "MEDIUM", "HIGH", "YouTube tutorials, influencer partnerships, organic growth"),
        ("Server costs explode", "MEDIUM", "MEDIUM", "CDN + edge cache, lazy load data, cost monitoring"),
        ("Competitor (Google Maps) copies", "HIGH", "HIGH", "Move faster, build community lock-in, data moat"),
        ("Key person risk (Arnaud)", "HIGH", "CRITICAL", "Document everything, hire co-founder ASAP, knowledge transfer"),
        ("Legal liability prospection", "MEDIUM", "CRITICAL", "Zero guarantees, user responsible, get insurance, consult lawyer"),
    ]
    
    row = 2
    for risk, prob, impact, mitigation in risks_data:
        ws_risks.cell(row=row, column=1).value = risk
        ws_risks.cell(row=row, column=2).value = prob
        ws_risks.cell(row=row, column=3).value = impact
        ws_risks.cell(row=row, column=4).value = mitigation
        row += 1
    
    # ========================================================================
    # SHEET 21 : MONETIZATION DETAILED
    # ========================================================================
    
    ws_monetization = wb.create_sheet("21_MONETIZATION", 21)
    for col in range(1, 7):
        ws_monetization.column_dimensions[get_column_letter(col)].width = 18
    
    apply_header(ws_monetization, 1, ["PRODUCT", "TIMING", "PRICE", "VOLUME EST.", "REVENUE Y1", "MARGIN"])
    
    monetization_data = [
        ("WalkSense Free", "W1", "0€", "500 users", "0€", "—"),
        ("WalkSense Premium", "W24", "5€/mois", "50 users avg", "3k€", "100%"),
        ("RockSense Kit V16", "W32", "500€", "10 units", "5k€", "10%"),
        ("MapSense Data API", "W44", "1000€/an", "2 orgs", "2k€", "80%"),
        ("Premium Analytics", "W50", "15€/mois", "20 users", "3.6k€", "95%"),
        ("Group Collaboration", "W40", "2€/mois premium", "30 users", "720€", "95%"),
        ("TOTAL REVENUE Y1", "—", "—", "~600 users", "~14.3k€", "—"),
    ]
    
    row = 2
    for product, timing, price, volume, revenue, margin in monetization_data:
        ws_monetization.cell(row=row, column=1).value = product
        ws_monetization.cell(row=row, column=2).value = timing
        ws_monetization.cell(row=row, column=3).value = price
        ws_monetization.cell(row=row, column=4).value = volume
        ws_monetization.cell(row=row, column=5).value = revenue
        ws_monetization.cell(row=row, column=6).value = margin
        row += 1
    
    # ========================================================================
    # SHEET 22 : TEAM & RESOURCES
    # ========================================================================
    
    ws_team = wb.create_sheet("22_TEAM_RESOURCES", 22)
    ws_team.column_dimensions['A'].width = 30
    ws_team.column_dimensions['B'].width = 20
    ws_team.column_dimensions['C'].width = 15
    ws_team.column_dimensions['D'].width = 25
    
    apply_header(ws_team, 1, ["RÔLE", "AFFECTÉ À", "STATUS", "CAPACITÉ"])
    
    team_data = [
        ("Product Manager", "Arnaud N.", "🔴 Solo", "Décisions produit, priorités, roadmap"),
        ("CTO / Lead Dev", "Arnaud N.", "🔴 Solo", "Architecture, code review, decisions tech"),
        ("Full-Stack Dev", "Arnaud N.", "🔴 Solo", "Mobile + Backend (W44+) development"),
        ("UI/UX Designer", "À EMBAUCHER", "🟡 W24", "Mockups, design system, user testing"),
        ("Hardware Engineer", "À EMBAUCHER", "🟡 W32", "RockSense PCB, firmware, testing"),
        ("DevOps / Infra", "À EMBAUCHER", "🟡 W44", "CI/CD, AWS/Hetzner setup, monitoring"),
        ("QA / Test", "À EMBAUCHER", "🟡 W25", "Test automation, field testing"),
        ("Sales / Growth", "À EMBAUCHER", "🟡 W40", "Marketing, user acquisition, partnerships"),
    ]
    
    row = 2
    for role, affecte, status, capacite in team_data:
        ws_team.cell(row=row, column=1).value = role
        ws_team.cell(row=row, column=2).value = affecte
        ws_team.cell(row=row, column=3).value = status
        ws_team.cell(row=row, column=4).value = capacite
        row += 1
    
    # ========================================================================
    # SHEET 23 : ACQUISITION & COMMUNITY
    # ========================================================================
    
    ws_acq = wb.create_sheet("23_ACQUISITION", 23)
    ws_acq.column_dimensions['A'].width = 30
    ws_acq.column_dimensions['B'].width = 70
    
    apply_header(ws_acq, 1, ["CHANNEL", "DESCRIPTION"])
    
    acq_data = [
        ("Associations prospecteurs", "Contact clubs locaux, FFMD, associations terrain"),
        ("Groupes Facebook", "Prospecteurs, archéologues, explorateurs, randonneurs"),
        ("YouTube", "Tutoriels démo, comparaison Nokta, prospection française"),
        ("Boutiques spécialisées", "Revendeurs détecteurs, partenariat affiliate"),
        ("Influenceurs YouTube", "Mini-sponsorship, unboxing, review RockSense"),
        ("Forums terrain", "MD-Forum, DetectoristForum, forums archéo français"),
        ("Reddit & Discord", "r/metaldetecting, serveurs Discord prospecteurs"),
        ("Campagne Google Ads", "€500/mois (W25+), ciblage Nokta + detection keywords"),
        ("Bouche-à-oreille", "Beta testers, early adopters, références organiques"),
        ("Événements", "Salons détecteurs, expo archéo, rassemblements prospecteurs"),
    ]
    
    row = 2
    for channel, description in acq_data:
        ws_acq.cell(row=row, column=1).value = channel
        ws_acq.cell(row=row, column=2).value = description
        row += 1
    
    # ========================================================================
    # SHEET 24 : NEXT STEPS IMMÉDIATS
    # ========================================================================
    
    ws_next = wb.create_sheet("24_NEXT_STEPS", 24)
    ws_next.column_dimensions['A'].width = 20
    ws_next.column_dimensions['B'].width = 70
    
    apply_header(ws_next, 1, ["PRIORITY", "ACTION"])
    
    next_steps_data = [
        ("🔴 THIS WEEK", "Finish Sprint A.1: TypeScript fixes + unit tests → Ready for field test"),
        ("🔴 W9-W10", "Field test V1.0: 30min session, 100m+ distance → Collect bugs + feedback"),
        ("🔴 W11-W12", "Fix identified bugs: GPS drift, photo orientation, map lag → Sprint A.2"),
        ("🟠 W13-W16", "Sprint B.1: Enhance features (photos scale, export perfection, notes UI)"),
        ("🟠 W17-W20", "Beta test with 5-10 real prospectors → Iterate on feedback"),
        ("🟠 W21-W24", "UI/UX redesign (hire designer W24) → Polish every pixel"),
        ("🟡 W25-W28", "Launch public alpha on Play Store (WalkSense v1.0-alpha.1)"),
        ("🟡 W29-W32", "Group mode implementation + hiring backend developer"),
        ("🟡 W32 DECISION", "Decision checkpoint: Build RockSense hardware or stay software-only?"),
        ("🟢 W33-W40", "If Hardware: RockSense proto + firmware. If Software: MapSense MVP."),
        ("🟢 W40+", "Public beta launch + marketing (YouTube, forums, influencers)"),
        ("🔵 W44+", "Series A fundraising (€200k) + backend scaling + cloud sync"),
    ]
    
    row = 2
    for priority, action in next_steps_data:
        ws_next.cell(row=row, column=1).value = priority
        ws_next.cell(row=row, column=2).value = action
        row += 1
    
    # ========================================================================
    # SHEET 25 : GLOSSARY
    # ========================================================================
    
    ws_glossary = wb.create_sheet("25_GLOSSARY", 25)
    ws_glossary.column_dimensions['A'].width = 25
    ws_glossary.column_dimensions['B'].width = 70
    
    apply_header(ws_glossary, 1, ["TERME", "DÉFINITION"])
    
    glossary_data = [
        ("WalkSense", "Application mobile de suivi GPS + documentation prospection"),
        ("RockSense", "Détecteur EM professionnel (future) avec capteurs géophysiques"),
        ("SeaSense", "Plateforme sonar marin (future)"),
        ("MapSense", "Moteur cartographique cloud + IA (future), fusion multicouche"),
        ("GeoSense", "Écosystème global = WalkSense + RockSense + SeaSense + MapSense"),
        ("Grid System", "Système de cellules spatiales (A1-P16) pour organiser zone prospection"),
        ("ERT", "Electrical Resistivity Tomography (tomographie résistivité électrique)"),
        ("IFT", "Indice de Fiabilité Terrain (score communautaire + comportemental)"),
        ("Geo-Fusion", "Moteur fusion GPS + IMU + BLE + WiFi RTT pour positionnement robuste"),
        ("HiddenMap", "Couche invisible documentaire : couverture GPS, densité, répétitions"),
        ("Offline-first", "Architecture où app fonctionne sans internet, sync après reconnexion"),
        ("Privacy-first", "Données utilisateur locales par défaut, opt-in cloud obligatoire"),
        ("Terrain-first", "Décisions guidées par réalités terrain, pas académiques"),
        ("Propriétaire", "Propriétaire du terrain où prospection a lieu (agriculteur, privé)"),
        ("Signature", "Consentement numérique propriétaire avant prospection"),
        ("Hash", "Empreinte numérique SHA-256 pour intégrité session immutable"),
        ("Freemium", "Modèle: app gratuite + premium features payantes (€5/mois)"),
        ("CAC", "Cost Acquisition Customer (coût acquisition utilisateur)"),
        ("LTV", "Lifetime Value (valeur durée vie client)"),
        ("Breakeven", "Point où revenus = coûts (projection Y2-Y3)"),
        ("Series A", "Financement tour A (€200k projection W44+)"),
        ("MVP", "Minimum Viable Product (produit minimum viable pour test marché)"),
        ("UI/UX", "User Interface / User Experience (interface + expérience utilisateur)"),
        ("BLE", "Bluetooth Low Energy (protocole sans fil basse consommation)"),
        ("GPS", "Global Positioning System (positionnement par satellite)"),
    ]
    
    row = 2
    for terme, definition in glossary_data:
        ws_glossary.cell(row=row, column=1).value = terme
        ws_glossary.cell(row=row, column=2).value = definition
        row += 1
    
    # ========================================================================
    # SHEET 26 : DOCX V7.3 SYNTHESIS
    # ========================================================================
    
    ws_docx_synthesis = wb.create_sheet("26_DOCX_SYNTHESIS", 26)
    ws_docx_synthesis.column_dimensions['A'].width = 35
    ws_docx_synthesis.column_dimensions['B'].width = 70
    
    title = ws_docx_synthesis['A1']
    title.value = "GeoSense WalkSense V7.3 MASTER CORE — Synthesis (DOCX intégré)"
    title.font = Font(bold=True, size=13, color="FFFFFF")
    title.fill = critical_fill
    ws_docx_synthesis.merge_cells('A1:B1')
    
    apply_header(ws_docx_synthesis, 2, ["PARTIE DOCX", "RÉSUMÉ INTÉGRÉ"])
    
    docx_synthesis = [
        ("PARTIE 1 — Vision Globale", "GeoSense fusionne documentation terrain + géospatial + cartographie + données + géophysique + communauté"),
        ("PARTIE 2 — Doctrine Produit", "Utile avant morale ✅ | Adoption avant monétisation ✅ | Simplicité avant sophistication ✅"),
        ("PARTIE 3 — Architecture Tech", "Expo 54 + React Native + TypeScript + SQLite + Expo Router + Expo Location + Crypto"),
        ("PARTIE 4 — Architecture SQLite", "Sessions | GPS Points | Photos | Markers | Overlays | Users | Groups | IFT Scores"),
        ("PARTIE 5 — Cloud Future", "Synchronisation | Sauvegarde | Historique | Partage | Dashboards | Analytics | API"),
        ("PARTIE 6 — WalkSense", "GPS live ✅ | Sessions ✅ | Historique ✅ | Replay ✅ | Photos ✅ | Marqueurs ✅ | Export ✅"),
        ("PARTIE 7 — MapSense", "Fusion multicouche : IGN + OSM + Satellite + Cassini + État-Major + Cadastre + LiDAR"),
        ("PARTIE 8 — Annexe MapSense", "Géoréférencement dynamique | Recalage | Cache offline | Tuiles | MNT | Ombrage relief"),
        ("PARTIE 9 — Geo-Fusion Engine", "Smoothing GPS | IMU fusion | Dead reckoning | RSSI BLE | WiFi RTT | Mesh collaboratif"),
        ("PARTIE 10 — Annexe Geo-Fusion", "Filtre Kalman | Inertie | Clustering | Recalage dynamique | Validation croisée"),
        ("PARTIE 11 — HiddenMap", "Couche invisible : GPS + densité couverture + répétitions + vitesse + heading + stabilité"),
        ("PARTIE 12 — IFT", "Indice Fiabilité Terrain : récompense cohérence + documentation + responsabilité + contribution"),
        ("PARTIE 13 — Annexe IFT", "Anti-abus : montée lente | validation | cohérence | historique | limitation spam"),
        ("PARTIE 14 — Groupes & Clubs", "Sorties collectives | Replay groupe | Dashboards | Sessions partagées | Clubs officiels"),
        ("PARTIE 15 — Accès Terrain", "Infrastructure confiance : propriétaires + agriculteurs + utilisateurs responsables"),
        ("PARTIE 16 — RockSense", "ESP32 + ADS1115 + INA826 | ERT | Cartographie sous-sol | Anomalies | Fusion surface/sous-sol"),
        ("PARTIE 17 — Annexe RockSense", "Électrodes | Acquisition | Inversion | Protocoles | Calibration | Reconstruction sous-sol"),
        ("PARTIE 18 — Roadmap Consolidée", "V1 (stabilité) → V1.5 (groupes) → V2 (cloud) → V2.5 (badges) → V3 (HiddenMap) → V4+ (API)"),
        ("PARTIE 19 — Acquisition", "Associations | Facebook | YouTube | Boutiques | Clubs terrain | Bouche-à-oreille organique"),
        ("PARTIE 20 — Business", "Freemium + Premium €5/mois + RockSense €500 + Data API €1000/an | Objectif €2.5-3k€/mois"),
        ("PARTIE 21 — Risques", "GPS imprécis ⚠️ | Hardware delayed | Saturation marché | RGPD backlash | Batterie drain"),
        ("PARTIE 22 — Directives IA", "Ne jamais simplifier ⚠️ | Enrichir ✅ | Consolider ✅ | Documenter ✅ | Préserver historique ✅"),
        ("PARTIE 23 — Analyse Finale", "Projet devient infrastructure documentaire + moteur géospatial + standard comportemental"),
    ]
    
    row = 3
    for partie, resume in docx_synthesis:
        ws_docx_synthesis.cell(row=row, column=1).value = partie
        ws_docx_synthesis.cell(row=row, column=2).value = resume
        row += 1
    
    # ========================================================================
    # SHEET 7 : SEASENSE VISION
    # ========================================================================
    
    ws_seasense = wb.create_sheet("07_SEASENSE_VISION", 7)
    ws_seasense.column_dimensions['A'].width = 35
    ws_seasense.column_dimensions['B'].width = 70
    
    title = ws_seasense['A1']
    title.value = "SeaSense — Sonar Marin (DOCX PARTIE ⚠️ À DÉFINIR)"
    title.font = Font(bold=True, size=12, color="FFFFFF")
    title.fill = critical_fill
    ws_seasense.merge_cells('A1:B1')
    
    apply_header(ws_seasense, 2, ["ASPECT", "STATUT / DÉFINITION"])
    
    seasense_data = [
        ("CONCEPT", "À DÉFINIR : Sonar portable ? Cartographie cours d'eau ? Scan littoral ?"),
        ("HARDWARE", "À DÉFINIR : Sonar écholoticateur ? Drone aquatique ? Capteur submersible ?"),
        ("SOFTWARE", "À DÉFINIR : Integration GPS / BLE / Data"),
        ("MARCHÉ", "À DÉFINIR : Archéologie aquatique ? Exploration littoral ? Pêche ? Sauvetage ?"),
        ("TIMELINE", "À DÉFINIR : Phase research uniquement (W24+)"),
        ("BUDGET", "À DÉFINIR : 5k-20k€ selon complexité"),
        ("PRIORITY", "🔵 FUTURE : Dépend succès WalkSense"),
        ("OWNER", "À ASSIGNER"),
    ]
    
    row = 3
    for aspect, value in seasense_data:
        ws_seasense.cell(row=row, column=1).value = aspect
        ws_seasense.cell(row=row, column=2).value = value
        row += 1
    
    # ========================================================================
    # SHEET 8 : MAPSENSE VISION
    # ========================================================================
    
    ws_mapsense = wb.create_sheet("08_MAPSENSE_VISION", 8)
    ws_mapsense.column_dimensions['A'].width = 35
    ws_mapsense.column_dimensions['B'].width = 70
    
    title = ws_mapsense['A1']
    title.value = "MapSense — Cloud + IA + Data (DOCX PARTIE 7-8)"
    title.font = Font(bold=True, size=12, color="FFFFFF")
    title.fill = critical_fill
    ws_mapsense.merge_cells('A1:B1')
    
    apply_header(ws_mapsense, 2, ["ASPECT", "DESCRIPTION"])
    
    mapsense_data = [
        ("VISION", "Plateforme cloud : fusion WalkSense + RockSense + SeaSense + IA"),
        ("OBJECTIF", "Base collective anonyme, patterns spatiaux, HiddenMap IA"),
        ("DATA COLLECTION", "Sessions WalkSense → optionnel upload cloud (user consent)"),
        ("ANONYMISATION", "Suppression identité utilisateur, conservation GPS/type/classe"),
        ("STORAGE", "PostgreSQL on Hetzner (EU) ou AWS (RGPD compliant)"),
        ("API", "REST API : GET /cells/{id}, POST /session, GET /heatmap/{region}"),
        ("VISUALIZATION", "Web dashboard : heatmap densité, timeline, filtres géo/type"),
        ("ML FUTURE", "IA classification image, prédiction zones riches, clustering spatial"),
        ("REVENUE", "B2B : data access pour archéologues/institutions (€1000/an)"),
        ("TIMELINE PHASE 1", "W44+ : Backend simple"),
        ("TIMELINE PHASE 2", "W60+ : ML avancé"),
        ("COMPLIANCE", "RGPD : opt-in, data anonymized, user can request deletion"),
    ]
    
    row = 3
    for aspect, description in mapsense_data:
        ws_mapsense.cell(row=row, column=1).value = aspect
        ws_mapsense.cell(row=row, column=2).value = description
        row += 1
    
    # ========================================================================
    # SHEET 9 : GRID SYSTEM
    # ========================================================================
    
    ws_grid = wb.create_sheet("09_GRID_SYSTEM", 9)
    ws_grid.column_dimensions['A'].width = 35
    ws_grid.column_dimensions['B'].width = 70
    
    title = ws_grid['A1']
    title.value = "Grid System — Cellules Spatiales (DOCX PARTIE 3 MASTER)"
    title.font = Font(bold=True, size=12, color="FFFFFF")
    title.fill = critical_fill
    ws_grid.merge_cells('A1:B1')
    
    apply_header(ws_grid, 2, ["ASPECT", "DESCRIPTION"])
    
    grid_data = [
        ("CONCEPT", "Core : 4 piquets = 4 points GPS connus → création grille virtuelle"),
        ("GRID DEFINITION", "Rectangle 4x4 ou 8x8 ou 16x16 cellules (configurable)"),
        ("CELLULE", "A1, A2, ..., P16 = adressable par prospecteur"),
        ("POSITION ESTIMATION", "GPS brut → projection sur grille → cellule estimée"),
        ("PRECISION", "Dépend : taille grille + précision GPS (±5m standard)"),
        ("MATH FORMULA", "Transformation affine : (lat,lon) → (grid_x, grid_y) → cellule"),
        ("IMPLEMENTATION", "Fonction simple trigonométrie en TypeScript/Kotlin"),
        ("USE CASE", "Prospecteur marque zone riche → peut dire 'C4, C5, D5' vs GPS exact"),
        ("AVANTAGE", "Communication simple, pas doublons zones, partage groupe facile"),
        ("UI DISPLAY", "Overlay grille sur carte + affichage cellule courante"),
        ("DATABASE", "Table cells : id, grid_id, name, markers_count, coverage %"),
        ("FUTURE", "Triangulation multi-piquet pour améliorer précision"),
    ]
    
    row = 3
    for aspect, description in grid_data:
        ws_grid.cell(row=row, column=1).value = aspect
        ws_grid.cell(row=row, column=2).value = description
        row += 1
    
    # ========================================================================
    # SHEET 10 : DATABASE SCHEMA (COMPLET)
    # ========================================================================
    
    ws_db = wb.create_sheet("10_DATABASE_SCHEMA", 10)
    for col in range(1, 8):
        ws_db.column_dimensions[get_column_letter(col)].width = 16
    
    apply_header(ws_db, 1, ["TABLE", "COLONNE", "TYPE", "PRIMAIRE", "CLÉS ÉTRANGÈRES", "INDEX", "DESCRIPTION"])
    
    db_schema = [
        ("users", "id", "String", "PK", "—", "yes", "UUID"),
        ("users", "name", "String", "—", "—", "—", "Nom utilisateur"),
        ("users", "email", "String", "UNIQUE", "—", "yes", "Email login"),
        ("users", "created_at", "Long", "—", "—", "—", "Timestamp création"),
        ("sessions", "id", "String", "PK", "—", "yes", "UUID"),
        ("sessions", "user_id", "String", "—", "users.id", "yes", "FK"),
        ("sessions", "title", "String", "—", "—", "—", "Nom session"),
        ("sessions", "start_time", "Long", "—", "—", "yes", "Timestamp début"),
        ("sessions", "end_time", "Long?", "—", "—", "—", "Timestamp fin"),
        ("sessions", "status", "Enum", "—", "—", "yes", "ONGOING/PAUSED/CLOSED"),
        ("sessions", "location", "String", "—", "—", "—", "Commune + département"),
        ("sessions", "grid_id", "String", "—", "grids.id", "—", "FK si grid"),
        ("sessions", "hash", "String", "—", "—", "yes", "SHA-256 intégrité"),
        ("sessions", "is_locked", "Boolean", "—", "—", "—", "Locked after close"),
        ("gps_points", "id", "String", "PK", "—", "yes", "UUID"),
        ("gps_points", "session_id", "String", "—", "sessions.id", "yes", "FK"),
        ("gps_points", "lat", "Double", "—", "—", "yes", "Latitude"),
        ("gps_points", "lon", "Double", "—", "—", "yes", "Longitude"),
        ("gps_points", "accuracy", "Float", "—", "—", "—", "Précision mètres"),
        ("gps_points", "timestamp", "Long", "—", "—", "yes", "Quand collecté"),
        ("gps_points", "speed", "Float?", "—", "—", "—", "Vitesse m/s"),
        ("markers", "id", "String", "PK", "—", "yes", "UUID"),
        ("markers", "session_id", "String", "—", "sessions.id", "yes", "FK"),
        ("markers", "lat", "Double", "—", "—", "yes", "Position"),
        ("markers", "lon", "Double", "—", "—", "yes", "Position"),
        ("markers", "classification", "Enum", "—", "—", "yes", "DECHET/MODERNE/..."),
        ("markers", "notes", "String?", "—", "—", "—", "Notes libres"),
        ("markers", "cell_id", "String", "—", "cells.id", "—", "FK si grid"),
        ("photos", "id", "String", "PK", "—", "yes", "UUID"),
        ("photos", "marker_id", "String", "—", "markers.id", "yes", "FK"),
        ("photos", "file_path", "String", "—", "—", "—", "Local path"),
        ("photos", "scale_reference", "String?", "—", "—", "—", "Échelle"),
        ("photos", "taken_at", "Long", "—", "—", "yes", "Timestamp"),
        ("grids", "id", "String", "PK", "—", "yes", "UUID"),
        ("grids", "session_id", "String", "—", "sessions.id", "yes", "FK"),
        ("grids", "anchor1_lat", "Double", "—", "—", "—", "Piquet 1"),
        ("grids", "anchor1_lon", "Double", "—", "—", "—", "Piquet 1"),
        ("grids", "anchor2_lat", "Double", "—", "—", "—", "Piquet 2"),
        ("grids", "anchor2_lon", "Double", "—", "—", "—", "Piquet 2"),
        ("grids", "anchor3_lat", "Double", "—", "—", "—", "Piquet 3"),
        ("grids", "anchor3_lon", "Double", "—", "—", "—", "Piquet 3"),
        ("grids", "anchor4_lat", "Double", "—", "—", "—", "Piquet 4"),
        ("grids", "anchor4_lon", "Double", "—", "—", "—", "Piquet 4"),
        ("grids", "grid_size", "Int", "—", "—", "—", "4/8/16 cellules côté"),
        ("grids", "created_at", "Long", "—", "—", "—", "Timestamp création"),
        ("cells", "id", "String", "PK", "—", "yes", "UUID"),
        ("cells", "grid_id", "String", "—", "grids.id", "yes", "FK"),
        ("cells", "name", "String", "—", "—", "yes", "A1, B2, ..."),
        ("cells", "row", "Int", "—", "—", "—", "Ligne 0-15"),
        ("cells", "col", "Int", "—", "—", "—", "Colonne 0-15"),
        ("cells", "markers_count", "Int", "—", "—", "—", "Nb marqueurs dedans"),
        ("cells", "coverage_percent", "Float", "—", "—", "—", "% couvert par session"),
        ("events", "id", "String", "PK", "—", "yes", "UUID"),
        ("events", "session_id", "String", "—", "sessions.id", "yes", "FK"),
        ("events", "type", "Enum", "—", "—", "yes", "MARKER/BLE_DETECT/..."),
        ("events", "lat", "Double?", "—", "—", "—", "Position événement"),
        ("events", "lon", "Double?", "—", "—", "—", "Position événement"),
        ("events", "description", "String", "—", "—", "—", "Détail"),
        ("events", "created_at", "Long", "—", "—", "yes", "Timestamp"),
        ("events", "data", "JSON?", "—", "—", "—", "Extra data (BLE, etc.)"),
        ("ble_devices", "id", "String", "PK", "—", "yes", "UUID"),
        ("ble_devices", "session_id", "String", "—", "sessions.id", "yes", "FK"),
        ("ble_devices", "device_name", "String", "—", "—", "—", "Nom RockSense"),
        ("ble_devices", "mac_address", "String", "UNIQUE", "—", "yes", "Adresse BLE"),
        ("ble_devices", "battery", "Int", "—", "—", "—", "% batterie"),
        ("ble_devices", "firmware_version", "String", "—", "—", "—", "Version fw"),
        ("ble_devices", "last_seen", "Long", "—", "—", "—", "Last sync timestamp"),
        ("session_exports", "id", "String", "PK", "—", "yes", "UUID"),
        ("session_exports", "session_id", "String", "—", "sessions.id", "yes", "FK"),
        ("session_exports", "format", "Enum", "—", "—", "—", "JSON/GPX/PDF"),
        ("session_exports", "file_path", "String", "—", "—", "—", "Chemin fichier"),
        ("session_exports", "created_at", "Long", "—", "—", "yes", "Timestamp export"),
        ("session_exports", "file_size_bytes", "Long", "—", "—", "—", "Taille fichier"),
    ]
    
    row = 2
    for table, colonne, type_, primary, fk, index, description in db_schema:
        ws_db.cell(row=row, column=1).value = table
        ws_db.cell(row=row, column=2).value = colonne
        ws_db.cell(row=row, column=3).value = type_
        ws_db.cell(row=row, column=4).value = primary
        ws_db.cell(row=row, column=5).value = fk
        ws_db.cell(row=row, column=6).value = index
        ws_db.cell(row=row, column=7).value = description
        row += 1
    
    # ========================================================================
    # SAVE FILE
    # ========================================================================
    
    wb.save(output_path)
    
    print("\n" + "="*90)
    print("✅ FICHIER EXCEL INTÉGRÉ CRÉÉ AVEC SUCCÈS !")
    print("="*90)
    print(f"\n📊 Nom du fichier : {output_path}")
    print(f"📋 Nombre de sheets : {len(wb.sheetnames)}")
    print(f"\n📑 Contenu (26 sheets complets) :")
    for i, sheet_name in enumerate(wb.sheetnames, 1):
        print(f"   {i:2d}. {sheet_name}")
    print("\n" + "="*90)
    print("✨ DOCX V7.3 MASTER CORE + Excel v2.0 fusionnés INTELLIGEMMENT !")
    print("   Toutes les couches : doctrine, product, tech, data, business, roadmap")
    print("="*90 + "\n")

# ============================================================================
# EXÉCUTION
# ============================================================================

if __name__ == "__main__":
    docx_file = "GeoSense_WalkSense_V7_3_MASTER_CORE_ANNEXES.docx"
    output_file = "GeoSense_Ecosystem_Master_v2.1_INTEGRATED_COMPLET.xlsx"
    
    create_integrated_excel(docx_file, output_file)
